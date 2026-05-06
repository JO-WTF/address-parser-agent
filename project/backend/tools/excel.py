from __future__ import annotations

from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


class ExcelTool:
    """Excel read and write helper."""

    def read_headers(self, file_path: str) -> list[str]:
        df = pd.read_excel(file_path, nrows=0)
        return [str(c) for c in df.columns.tolist()]

    def read_column(self, file_path: str, column: str) -> Iterable[tuple[int, str]]:
        df = pd.read_excel(file_path, usecols=[column])
        for idx, value in enumerate(df[column].fillna(""), start=2):
            yield idx, str(value)

    def get_total_rows(self, file_path: str) -> int:
        wb = load_workbook(file_path, read_only=True)
        ws = wb.active
        total = max(ws.max_row - 1, 0)
        wb.close()
        return total

    def copy_to_output(self, file_path: str, output_path: str) -> None:
        Path(output_path).write_bytes(Path(file_path).read_bytes())

    def write_result_row(self, output_path: str, row_num: int, result: dict) -> None:
        wb = load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        new_columns = ["name", "phone", "address", "province", "city"]

        for col in new_columns:
            if col not in headers:
                ws.cell(row=1, column=len(headers) + 1, value=col)
                headers.append(col)

        for col in new_columns:
            col_idx = headers.index(col) + 1
            ws.cell(row=row_num, column=col_idx, value=result.get(col, ""))

        wb.save(output_path)
        wb.close()
